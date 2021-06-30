import openpyxl as xl
from openpyxl.chart import BarChart 

def process_book(filename):
    # Loading the excel in the file & stored in an object!
    workbook_obj = xl.load_workbook(filename)

    # Accessing the sheet in the excel file(sheet name is case sensitive):
    sheet = workbook_obj['Sheet1']

    # Accessing the cells in the sheet (2 ways):
    cell = sheet['a1']
    cell = sheet.cell(1,1)



    # (here we have total 4 rows)
    for row in range(2,sheet.max_row + 1):
        cell = sheet.cell(row,3)
        correct_price = cell.value * 0.9
        correct_price_cell = sheet.cell(row,4)
        correct_price_cell.value = correct_price

    # Using the Reference class to select the values
    values = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4)

    # Passing our values to the BarChart Class
    chart = BarChart()
    chart.add_data(values)

    # Adding the chart to our sheet:
    sheet.add_chart(chart, 'f2')


    # Saving the excel in a new file:
    workbook_obj.save(filename)

input_chart = input("Enter the name: ")

process_book(input_chart)



           

 

















